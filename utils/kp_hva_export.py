import tempfile
import os
import copy
import logging
from datetime import datetime
from typing import Dict, Any, Optional

logger = logging.getLogger(__name__)

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

_TEMPLATE_CANDIDATES = [
    os.path.join(_PROJECT_ROOT, 'data', 'templates', 'kp_hva_template.xlsm'),
    os.path.join(_PROJECT_ROOT, 'attached_assets', 'kp_incident_log_hva_(5)_1771953905117.xlsm'),
]

KP_TEMPLATE_PATH = None
for _candidate in _TEMPLATE_CANDIDATES:
    if os.path.exists(_candidate):
        KP_TEMPLATE_PATH = _candidate
        break

if KP_TEMPLATE_PATH is None:
    logger.warning(f"KP HVA template not found in any expected location: {_TEMPLATE_CANDIDATES}")
    KP_TEMPLATE_PATH = _TEMPLATE_CANDIDATES[0]

KP_HAZARD_ROW_MAP = {
    'Active Shooter': 14,
    'Air Quality Issue': 16,
    'Dam Failure': 25,
    'Drought': 26,
    'Epidemic': 28,
    'Flood, External': 33,
    'Inclement Weather': 41,
    'Infectious Disease Outbreak': 42,
    'IT System Outage': 43,
    'Pandemic': 50,
    'Power Outage': 55,
    'Supply Chain Shortage / Failure': 62,
    'Temperature Extremes': 64,
    'Tornado': 65,
    'Utility Failure': 69,
    'Water Contamination': 71,
    'Water Disruption': 72,
    'Communication / Telephony Failure': 24,
    'Seasonal Influenza': 57,
    'Civil Unrest / Protesting': 23,
}


def _score_to_kp_scale(cara_score: float) -> int:
    if cara_score is None or cara_score <= 0.0:
        return 0
    elif cara_score <= 0.33:
        return 1
    elif cara_score <= 0.66:
        return 2
    else:
        return 3


def _damage_to_kp_scale(damage_amount: float, hazard_type: str) -> int:
    thresholds = {
        'flood': (100_000, 5_000_000),
        'tornado': (50_000, 2_000_000),
        'winter_storm': (50_000, 1_000_000),
        'thunderstorm': (25_000, 500_000),
    }
    low_thresh, high_thresh = thresholds.get(hazard_type, (50_000, 2_000_000))

    if damage_amount <= 0:
        return 0
    elif damage_amount <= low_thresh:
        return 1
    elif damage_amount <= high_thresh:
        return 2
    else:
        return 3


def _injuries_fatalities_to_kp_scale(injuries: int, fatalities: int) -> int:
    if fatalities >= 5:
        return 3
    elif fatalities >= 1 or injuries >= 20:
        return 3
    elif injuries >= 5:
        return 2
    elif injuries >= 1:
        return 1
    else:
        return 0


def _health_factor_to_kp_scale(health_factor: float) -> int:
    if health_factor >= 1.4:
        return 3
    elif health_factor >= 1.2:
        return 2
    elif health_factor >= 1.0:
        return 1
    else:
        return 0


def _get_storm_raw_data(county_name: str) -> Dict[str, Dict[str, Any]]:
    try:
        from utils.noaa_storm_events import get_county_storm_summary
        storm_data = get_county_storm_summary(county_name)
        if not storm_data:
            return {}

        by_category = storm_data.get('by_category', {})
        result = {}
        for cat_key in ['flood', 'tornado', 'winter_storm', 'thunderstorm']:
            cat_data = by_category.get(cat_key, {})
            if cat_data:
                result[cat_key] = {
                    'event_count': cat_data.get('event_count', 0),
                    'injuries': cat_data.get('injuries', 0),
                    'fatalities': cat_data.get('fatalities', 0),
                    'property_damage': cat_data.get('property_damage', 0),
                }
        return result
    except Exception as e:
        logger.warning(f"Could not get NOAA storm data for {county_name}: {e}")
        return {}


def _get_openfema_raw_data(county_name: str) -> Dict[str, Any]:
    try:
        from utils.openfema_data import get_county_openfema_summary
        data = get_county_openfema_summary(county_name)
        return data if data else {}
    except Exception as e:
        logger.warning(f"Could not get OpenFEMA data for {county_name}: {e}")
        return {}


def _build_kp_hazard_data(risk_data: dict, county_name: str) -> Dict[str, Dict[str, int]]:
    storm_raw = _get_storm_raw_data(county_name)
    openfema_raw = _get_openfema_raw_data(county_name)

    from utils.risk_calculation import get_health_impact_factor

    hazard_data = {}

    flood_storm = storm_raw.get('flood', {})
    flood_risk = risk_data.get('flood_risk', 0.0)
    flood_hif = get_health_impact_factor(county_name, 'flood')
    nfip = openfema_raw.get('nfip_claims', {})
    nfip_claims = nfip.get('total_claims', 0) if nfip else 0

    flood_human = _injuries_fatalities_to_kp_scale(
        flood_storm.get('injuries', 0), flood_storm.get('fatalities', 0)
    )
    if flood_human == 0 and flood_hif >= 1.1:
        flood_human = _health_factor_to_kp_scale(flood_hif)

    flood_prop_raw = flood_storm.get('property_damage', 0)
    if nfip_claims > 50:
        flood_prop_raw = max(flood_prop_raw, 500_000)
    flood_property = _damage_to_kp_scale(flood_prop_raw, 'flood')

    hazard_data['Flood, External'] = {
        'probability': _score_to_kp_scale(flood_risk),
        'human_impact': flood_human,
        'property_impact': flood_property,
    }

    tornado_storm = storm_raw.get('tornado', {})
    tornado_risk = risk_data.get('tornado_risk', 0.0)
    tornado_hif = get_health_impact_factor(county_name, 'tornado')
    tornado_human = _injuries_fatalities_to_kp_scale(
        tornado_storm.get('injuries', 0), tornado_storm.get('fatalities', 0)
    )
    if tornado_human == 0:
        tornado_human = max(1, _health_factor_to_kp_scale(tornado_hif))
    tornado_property = _damage_to_kp_scale(
        tornado_storm.get('property_damage', 0), 'tornado'
    )
    if tornado_property == 0 and tornado_storm.get('event_count', 0) > 0:
        tornado_property = 1

    hazard_data['Tornado'] = {
        'probability': _score_to_kp_scale(tornado_risk),
        'human_impact': tornado_human,
        'property_impact': tornado_property,
    }

    winter_storm = storm_raw.get('winter_storm', {})
    winter_risk = risk_data.get('winter_storm_risk', 0.0)
    winter_hif = get_health_impact_factor(county_name, 'winter_storm')
    winter_human = _injuries_fatalities_to_kp_scale(
        winter_storm.get('injuries', 0), winter_storm.get('fatalities', 0)
    )
    if winter_human == 0:
        winter_human = _health_factor_to_kp_scale(winter_hif)
    winter_property = _damage_to_kp_scale(
        winter_storm.get('property_damage', 0), 'winter_storm'
    )

    hazard_data['Inclement Weather'] = {
        'probability': _score_to_kp_scale(winter_risk),
        'human_impact': winter_human,
        'property_impact': winter_property,
    }

    ts_storm = storm_raw.get('thunderstorm', {})
    ts_risk = risk_data.get('thunderstorm_risk', 0.0)

    dam_failure_risk = risk_data.get('dam_failure_risk', 0.0)
    if dam_failure_risk > 0:
        dam_prob = _score_to_kp_scale(dam_failure_risk)
        dam_human = _score_to_kp_scale(dam_failure_risk * 1.2)
        dam_property = _score_to_kp_scale(dam_failure_risk * 1.1)
    else:
        dam_prob = _score_to_kp_scale(flood_risk * 0.6)
        dam_human = min(3, flood_human + 1) if flood_human > 0 else 1
        dam_property = min(3, flood_property + 1) if flood_property > 0 else 1

    hazard_data['Dam Failure'] = {
        'probability': dam_prob,
        'human_impact': dam_human,
        'property_impact': dam_property,
    }

    active_shooter_risk = risk_data.get('active_shooter_risk', 0.0)
    hazard_data['Active Shooter'] = {
        'probability': _score_to_kp_scale(active_shooter_risk),
        'human_impact': 3,
        'property_impact': 1,
    }

    air_quality_risk = risk_data.get('air_quality_risk', 0.0)
    hazard_data['Air Quality Issue'] = {
        'probability': _score_to_kp_scale(air_quality_risk),
        'human_impact': _score_to_kp_scale(air_quality_risk),
        'property_impact': 0,
    }

    health_risk = risk_data.get('health_risk', 0.0)
    vector_borne_risk = risk_data.get('vector_borne_disease_risk', 0.0)
    combined_disease_risk = max(health_risk, vector_borne_risk)

    hazard_data['Epidemic'] = {
        'probability': _score_to_kp_scale(max(health_risk * 0.7, vector_borne_risk * 0.8)),
        'human_impact': _score_to_kp_scale(combined_disease_risk),
        'property_impact': 0,
    }
    hazard_data['Pandemic'] = {
        'probability': _score_to_kp_scale(health_risk),
        'human_impact': 3,
        'property_impact': 1,
    }
    hazard_data['Infectious Disease Outbreak'] = {
        'probability': _score_to_kp_scale(max(health_risk * 0.8, vector_borne_risk)),
        'human_impact': _score_to_kp_scale(combined_disease_risk),
        'property_impact': 0,
    }
    hazard_data['Seasonal Influenza'] = {
        'probability': 2,
        'human_impact': 1,
        'property_impact': 0,
    }

    extreme_heat_risk = risk_data.get('extreme_heat_risk', 0.0)
    heat_hif = get_health_impact_factor(county_name, 'extreme_heat')
    hazard_data['Temperature Extremes'] = {
        'probability': _score_to_kp_scale(extreme_heat_risk),
        'human_impact': max(1, _health_factor_to_kp_scale(heat_hif)),
        'property_impact': 1,
    }

    cyber_risk = risk_data.get('cybersecurity_risk', 0.0)
    hazard_data['IT System Outage'] = {
        'probability': _score_to_kp_scale(cyber_risk),
        'human_impact': 1,
        'property_impact': _score_to_kp_scale(cyber_risk),
    }
    hazard_data['Communication / Telephony Failure'] = {
        'probability': _score_to_kp_scale(cyber_risk * 0.7),
        'human_impact': 1,
        'property_impact': _score_to_kp_scale(cyber_risk * 0.5),
    }

    utilities_data = risk_data.get('utilities', {})
    electrical_risk = utilities_data.get('electrical_outage', 0.5)
    utilities_overall = utilities_data.get('overall', 0.5)
    supply_chain_risk = utilities_data.get('supply_chain', 0.5)

    hazard_data['Power Outage'] = {
        'probability': _score_to_kp_scale(electrical_risk),
        'human_impact': 1,
        'property_impact': _score_to_kp_scale(electrical_risk * 0.6),
    }
    hazard_data['Utility Failure'] = {
        'probability': _score_to_kp_scale(utilities_overall),
        'human_impact': 1,
        'property_impact': _score_to_kp_scale(utilities_overall * 0.7),
    }
    hazard_data['Supply Chain Shortage / Failure'] = {
        'probability': _score_to_kp_scale(supply_chain_risk),
        'human_impact': 1,
        'property_impact': _score_to_kp_scale(supply_chain_risk * 0.5),
    }

    hazard_data['Water Contamination'] = {
        'probability': 1,
        'human_impact': 2,
        'property_impact': 1,
    }
    hazard_data['Water Disruption'] = {
        'probability': 1,
        'human_impact': 1,
        'property_impact': 1,
    }

    hazard_data['Drought'] = {
        'probability': _score_to_kp_scale(extreme_heat_risk * 0.5),
        'human_impact': 1,
        'property_impact': 1,
    }

    hazard_data['Civil Unrest / Protesting'] = {
        'probability': _score_to_kp_scale(active_shooter_risk * 0.4),
        'human_impact': 1,
        'property_impact': 1,
    }

    return hazard_data


def _aggregate_storm_data_for_counties(counties: list) -> Dict[str, Dict[str, Any]]:
    aggregated = {}
    for county in counties:
        county_storm = _get_storm_raw_data(county)
        for cat_key, cat_data in county_storm.items():
            if cat_key not in aggregated:
                aggregated[cat_key] = {'event_count': 0, 'injuries': 0, 'fatalities': 0, 'property_damage': 0}
            aggregated[cat_key]['event_count'] += cat_data.get('event_count', 0)
            aggregated[cat_key]['injuries'] += cat_data.get('injuries', 0)
            aggregated[cat_key]['fatalities'] += cat_data.get('fatalities', 0)
            aggregated[cat_key]['property_damage'] += cat_data.get('property_damage', 0)
    return aggregated


def _aggregate_openfema_data_for_counties(counties: list) -> Dict[str, Any]:
    total_nfip_claims = 0
    total_nfip_payout = 0
    total_flood_decl = 0
    has_nfip = False
    has_decl = False

    for county in counties:
        fema = _get_openfema_raw_data(county)
        nfip = fema.get('nfip_claims')
        if nfip:
            has_nfip = True
            total_nfip_claims += nfip.get('total_claims', 0)
            total_nfip_payout += nfip.get('total_payout', 0)
        decl = fema.get('disaster_declarations')
        if decl:
            has_decl = True
            total_flood_decl += decl.get('by_incident_type', {}).get('Flood', 0)

    result = {}
    if has_nfip:
        result['nfip_claims'] = {'total_claims': total_nfip_claims, 'total_payout': total_nfip_payout}
    if has_decl:
        result['disaster_declarations'] = {'by_incident_type': {'Flood': total_flood_decl}}
    return result


def _avg_health_impact_factor(counties: list, hazard_type: str) -> float:
    from utils.risk_calculation import get_health_impact_factor
    factors = []
    for county in counties:
        try:
            factors.append(get_health_impact_factor(county, hazard_type))
        except Exception as e:
            logger.debug(f"Could not get health impact factor for {county}/{hazard_type}: {e}")
    return sum(factors) / len(factors) if factors else 1.0


def _build_kp_hazard_data_herc(risk_data: dict, counties: list) -> Dict[str, Dict[str, int]]:
    storm_raw = _aggregate_storm_data_for_counties(counties)
    openfema_raw = _aggregate_openfema_data_for_counties(counties)

    hazard_data = {}

    flood_storm = storm_raw.get('flood', {})
    flood_risk = risk_data.get('flood_risk', 0.0)
    flood_hif = _avg_health_impact_factor(counties, 'flood')
    nfip = openfema_raw.get('nfip_claims', {})
    nfip_claims = nfip.get('total_claims', 0) if nfip else 0

    flood_human = _injuries_fatalities_to_kp_scale(
        flood_storm.get('injuries', 0), flood_storm.get('fatalities', 0)
    )
    if flood_human == 0 and flood_hif >= 1.1:
        flood_human = _health_factor_to_kp_scale(flood_hif)

    flood_prop_raw = flood_storm.get('property_damage', 0)
    if nfip_claims > 50:
        flood_prop_raw = max(flood_prop_raw, 500_000)
    flood_property = _damage_to_kp_scale(flood_prop_raw, 'flood')

    hazard_data['Flood, External'] = {
        'probability': _score_to_kp_scale(flood_risk),
        'human_impact': flood_human,
        'property_impact': flood_property,
    }

    tornado_storm = storm_raw.get('tornado', {})
    tornado_risk = risk_data.get('tornado_risk', 0.0)
    tornado_hif = _avg_health_impact_factor(counties, 'tornado')
    tornado_human = _injuries_fatalities_to_kp_scale(
        tornado_storm.get('injuries', 0), tornado_storm.get('fatalities', 0)
    )
    if tornado_human == 0:
        tornado_human = max(1, _health_factor_to_kp_scale(tornado_hif))
    tornado_property = _damage_to_kp_scale(
        tornado_storm.get('property_damage', 0), 'tornado'
    )
    if tornado_property == 0 and tornado_storm.get('event_count', 0) > 0:
        tornado_property = 1

    hazard_data['Tornado'] = {
        'probability': _score_to_kp_scale(tornado_risk),
        'human_impact': tornado_human,
        'property_impact': tornado_property,
    }

    winter_storm = storm_raw.get('winter_storm', {})
    winter_risk = risk_data.get('winter_storm_risk', 0.0)
    winter_hif = _avg_health_impact_factor(counties, 'winter_storm')
    winter_human = _injuries_fatalities_to_kp_scale(
        winter_storm.get('injuries', 0), winter_storm.get('fatalities', 0)
    )
    if winter_human == 0:
        winter_human = _health_factor_to_kp_scale(winter_hif)
    winter_property = _damage_to_kp_scale(
        winter_storm.get('property_damage', 0), 'winter_storm'
    )

    hazard_data['Inclement Weather'] = {
        'probability': _score_to_kp_scale(winter_risk),
        'human_impact': winter_human,
        'property_impact': winter_property,
    }

    dam_failure_risk = risk_data.get('dam_failure_risk', 0.0)
    if dam_failure_risk > 0:
        dam_prob = _score_to_kp_scale(dam_failure_risk)
        dam_human = _score_to_kp_scale(dam_failure_risk * 1.2)
        dam_property = _score_to_kp_scale(dam_failure_risk * 1.1)
    else:
        dam_prob = _score_to_kp_scale(flood_risk * 0.6)
        dam_human = min(3, flood_human + 1) if flood_human > 0 else 1
        dam_property = min(3, flood_property + 1) if flood_property > 0 else 1

    hazard_data['Dam Failure'] = {
        'probability': dam_prob,
        'human_impact': dam_human,
        'property_impact': dam_property,
    }

    active_shooter_risk = risk_data.get('active_shooter_risk', 0.0)
    hazard_data['Active Shooter'] = {
        'probability': _score_to_kp_scale(active_shooter_risk),
        'human_impact': 3,
        'property_impact': 1,
    }

    air_quality_risk = risk_data.get('air_quality_risk', 0.0)
    hazard_data['Air Quality Issue'] = {
        'probability': _score_to_kp_scale(air_quality_risk),
        'human_impact': _score_to_kp_scale(air_quality_risk),
        'property_impact': 0,
    }

    health_risk = risk_data.get('health_risk', 0.0)
    vector_borne_risk = risk_data.get('vector_borne_disease_risk', 0.0)
    combined_disease_risk = max(health_risk, vector_borne_risk)

    hazard_data['Epidemic'] = {
        'probability': _score_to_kp_scale(max(health_risk * 0.7, vector_borne_risk * 0.8)),
        'human_impact': _score_to_kp_scale(combined_disease_risk),
        'property_impact': 0,
    }
    hazard_data['Pandemic'] = {
        'probability': _score_to_kp_scale(health_risk),
        'human_impact': 3,
        'property_impact': 1,
    }
    hazard_data['Infectious Disease Outbreak'] = {
        'probability': _score_to_kp_scale(max(health_risk * 0.8, vector_borne_risk)),
        'human_impact': _score_to_kp_scale(combined_disease_risk),
        'property_impact': 0,
    }
    hazard_data['Seasonal Influenza'] = {
        'probability': 2,
        'human_impact': 1,
        'property_impact': 0,
    }

    extreme_heat_risk = risk_data.get('extreme_heat_risk', 0.0)
    heat_hif = _avg_health_impact_factor(counties, 'extreme_heat')
    hazard_data['Temperature Extremes'] = {
        'probability': _score_to_kp_scale(extreme_heat_risk),
        'human_impact': max(1, _health_factor_to_kp_scale(heat_hif)),
        'property_impact': 1,
    }

    cyber_risk = risk_data.get('cybersecurity_risk', 0.0)
    hazard_data['IT System Outage'] = {
        'probability': _score_to_kp_scale(cyber_risk),
        'human_impact': 1,
        'property_impact': _score_to_kp_scale(cyber_risk),
    }
    hazard_data['Communication / Telephony Failure'] = {
        'probability': _score_to_kp_scale(cyber_risk * 0.7),
        'human_impact': 1,
        'property_impact': _score_to_kp_scale(cyber_risk * 0.5),
    }

    utilities_risk = risk_data.get('utilities_risk', 0.5)
    hazard_data['Power Outage'] = {
        'probability': _score_to_kp_scale(utilities_risk),
        'human_impact': 1,
        'property_impact': _score_to_kp_scale(utilities_risk * 0.6),
    }
    hazard_data['Utility Failure'] = {
        'probability': _score_to_kp_scale(utilities_risk),
        'human_impact': 1,
        'property_impact': _score_to_kp_scale(utilities_risk * 0.7),
    }
    hazard_data['Supply Chain Shortage / Failure'] = {
        'probability': _score_to_kp_scale(utilities_risk * 0.7),
        'human_impact': 1,
        'property_impact': _score_to_kp_scale(utilities_risk * 0.5),
    }

    hazard_data['Water Contamination'] = {
        'probability': 1,
        'human_impact': 2,
        'property_impact': 1,
    }
    hazard_data['Water Disruption'] = {
        'probability': 1,
        'human_impact': 1,
        'property_impact': 1,
    }

    hazard_data['Drought'] = {
        'probability': _score_to_kp_scale(extreme_heat_risk * 0.5),
        'human_impact': 1,
        'property_impact': 1,
    }

    hazard_data['Civil Unrest / Protesting'] = {
        'probability': _score_to_kp_scale(active_shooter_risk * 0.4),
        'human_impact': 1,
        'property_impact': 1,
    }

    return hazard_data


def _write_kp_template(wb, entity_name: str, hazard_data: Dict[str, Dict[str, int]]) -> None:
    ws = wb['HVA']

    if 'Input' in wb.sheetnames:
        ws_input = wb['Input']
        ws_input['B3'] = entity_name
        ws_input['B4'] = f"CARA Auto-populated {datetime.now().strftime('%m/%d/%Y')}"

    if 'Reference' in wb.sheetnames:
        ws_ref = wb['Reference']
        ref_values = {}
        for ref_row in ws_ref.iter_rows(min_row=1, max_row=ws_ref.max_row, values_only=False):
            val = ref_row[0].value
            if val and isinstance(val, str):
                ref_values[ref_row[0].row] = val

        import re as _re
        for hva_row in ws.iter_rows(min_row=14, max_row=74, values_only=False):
            cell = hva_row[0]
            formula = cell.value
            if formula and isinstance(formula, str) and 'Reference!' in formula:
                m = _re.search(r'Reference!A(\d+)', formula)
                if m:
                    ref_row_num = int(m.group(1))
                    name = ref_values.get(ref_row_num)
                    if name:
                        cell.value = name

    populated_count = 0
    for hazard_name, row_num in KP_HAZARD_ROW_MAP.items():
        if hazard_name in hazard_data:
            hd = hazard_data[hazard_name]
            ws.cell(row=row_num, column=2, value=hd['probability'])
            ws.cell(row=row_num, column=5, value=hd['human_impact'])
            ws.cell(row=row_num, column=6, value=hd['property_impact'])
            populated_count += 1
            logger.debug(f"  Row {row_num} ({hazard_name}): B={hd['probability']}, E={hd['human_impact']}, F={hd['property_impact']}")

    logger.info(f"Populated {populated_count}/{len(KP_HAZARD_ROW_MAP)} hazards for {entity_name}")


def generate_kp_hva_herc_export(risk_data: dict) -> str:
    try:
        from openpyxl import load_workbook

        region_name = risk_data.get('name', risk_data.get('location', 'Unknown HERC Region'))
        counties = risk_data.get('counties', [])
        herc_id = risk_data.get('herc_id', 'unknown')

        logger.info(f"Generating KP HVA export for HERC {region_name} ({len(counties)} counties)")

        wb = load_workbook(KP_TEMPLATE_PATH, keep_vba=True)

        if 'HVA' not in wb.sheetnames:
            raise ValueError("KP template missing 'HVA' sheet")

        hazard_data = _build_kp_hazard_data_herc(risk_data, counties)
        _write_kp_template(wb, region_name, hazard_data)

        temp_dir = tempfile.gettempdir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = region_name.replace(' ', '_').replace('/', '-')
        filename = f'KP_HVA_HERC_{safe_name}_{timestamp}.xlsm'
        file_path = os.path.join(temp_dir, filename)

        wb.save(file_path)
        logger.info(f"Saved HERC KP HVA export to {file_path}")

        return file_path

    except Exception as e:
        logger.error(f"Error generating HERC KP HVA export: {e}")
        import traceback
        logger.debug(traceback.format_exc())
        raise Exception(f"Error generating Kaiser Permanente HVA export: {str(e)}")


def generate_kp_hva_jurisdiction_export(risk_data: dict) -> str:
    try:
        from openpyxl import load_workbook

        jurisdiction_name = risk_data.get('location', 'Unknown Jurisdiction')
        county_name = risk_data.get('county', risk_data.get('county_name', ''))

        logger.info(f"Generating KP HVA export for {jurisdiction_name} (county: {county_name})")

        wb = load_workbook(KP_TEMPLATE_PATH, keep_vba=True)

        if 'HVA' not in wb.sheetnames:
            raise ValueError("KP template missing 'HVA' sheet")

        hazard_data = _build_kp_hazard_data(risk_data, county_name)
        _write_kp_template(wb, jurisdiction_name, hazard_data)

        temp_dir = tempfile.gettempdir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = jurisdiction_name.replace(' ', '_').replace('/', '-')
        filename = f'KP_HVA_{safe_name}_{timestamp}.xlsm'
        file_path = os.path.join(temp_dir, filename)

        wb.save(file_path)
        logger.info(f"Saved KP HVA export to {file_path}")

        return file_path

    except Exception as e:
        logger.error(f"Error generating KP HVA export: {e}")
        import traceback
        logger.debug(traceback.format_exc())
        raise Exception(f"Error generating Kaiser Permanente HVA export: {str(e)}")
