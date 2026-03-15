import logging
import os
import json
import tempfile
import time
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
PH_FILL = PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid")
EM_FILL = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
DIFF_POS_FONT = Font(name="Calibri", size=10, color="c62828")
DIFF_NEG_FONT = Font(name="Calibri", size=10, color="2e7d32")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

DOMAINS = [
    ("total_risk_score", "Overall PHRAT Score"),
    ("natural_hazards_risk", "Natural Hazards"),
    ("flood_risk", "  Flood"),
    ("tornado_risk", "  Tornado"),
    ("winter_storm_risk", "  Winter Storm"),
    ("thunderstorm_risk", "  Thunderstorm"),
    ("health_risk", "Health Metrics"),
    ("active_shooter_risk", "Active Shooter"),
    ("air_quality_risk", "Air Quality"),
    ("extreme_heat_risk", "Extreme Heat"),
    ("dam_failure_risk", "Dam Failure"),
    ("vector_borne_disease_risk", "Vector-Borne Disease"),
    ("utilities_risk", "Utilities"),
]

SCORES_CACHE_KEY = "em_comparison_scores"
SCORES_MAX_AGE = 86400


def _get_cached_scores_from_db():
    try:
        from core import db
        from models import DataSourceCache
        from sqlalchemy import desc
        entry = db.session.query(DataSourceCache).filter(
            DataSourceCache.source_type == SCORES_CACHE_KEY,
            DataSourceCache.is_valid == True
        ).order_by(desc(DataSourceCache.fetched_at)).first()
        if entry:
            age = (datetime.utcnow() - entry.fetched_at).total_seconds()
            if age < SCORES_MAX_AGE:
                return entry.data
    except Exception as e:
        logger.warning(f"Error reading EM scores from DB cache: {e}")
    return None


def _save_scores_to_db(result):
    try:
        from core import db
        from models import DataSourceCache
        db.session.query(DataSourceCache).filter(
            DataSourceCache.source_type == SCORES_CACHE_KEY
        ).delete()
        entry = DataSourceCache(
            source_type=SCORES_CACHE_KEY,
            data=result,
            fetched_at=datetime.utcnow(),
            expires_at=datetime.utcnow() + timedelta(seconds=SCORES_MAX_AGE),
            is_valid=True,
            api_source="cara_internal"
        )
        db.session.add(entry)
        db.session.commit()
        logger.info(f"Saved EM comparison scores to DB cache ({result.get('count', 0)} jurisdictions)")
    except Exception as e:
        logger.error(f"Error saving EM scores to DB cache: {e}")
        db.session.rollback()


def precompute_comparison_scores():
    from utils.data_processor import get_em_jurisdictions, process_risk_data

    em_jurisdictions = get_em_jurisdictions()
    em_jurisdictions_sorted = sorted(em_jurisdictions, key=lambda x: x.get("em_name", x["name"]))

    rows = []
    total = len(em_jurisdictions_sorted)
    for idx, j in enumerate(em_jurisdictions_sorted):
        jid = j["id"]
        em_name = j.get("em_name", j["name"])

        try:
            ph_data = process_risk_data(jid, discipline="public_health")
            em_data = process_risk_data(jid, discipline="em")
        except Exception as e:
            logger.error(f"Error processing jurisdiction {jid} ({em_name}): {e}")
            continue

        if not ph_data or not em_data:
            continue

        ph_scores = {k: float(ph_data.get(k, 0) or 0) for k, _ in DOMAINS}
        em_scores = {k: float(em_data.get(k, 0) or 0) for k, _ in DOMAINS}
        rows.append({"name": em_name, "ph": ph_scores, "em": em_scores})

        if (idx + 1) % 10 == 0:
            logger.info(f"Pre-compute: {idx + 1}/{total} jurisdictions done")

    result = {
        "generated": datetime.now().isoformat(),
        "count": len(rows),
        "rows": rows,
    }

    _save_scores_to_db(result)
    logger.info(f"Pre-computed {len(rows)} jurisdiction scores")
    return result


def load_precomputed_scores():
    return _get_cached_scores_from_db()


def generate_em_comparison_export():
    scores_data = load_precomputed_scores()
    if not scores_data:
        raise RuntimeError(
            "Comparison scores have not been pre-computed yet. "
            "Please try again in a few minutes while the system generates the data."
        )

    rows = scores_data["rows"]
    generated = scores_data.get("generated", "Unknown")

    wb = Workbook()
    ws = wb.active
    ws.title = "PH vs EM Comparison"

    ws.merge_cells("A1:Z1")
    title_cell = ws["A1"]
    title_cell.value = "CARA Risk Score Comparison: Public Health vs Emergency Management Weights"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1a237e")
    title_cell.alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:Z2")
    ws["A2"].value = f"Scores computed: {generated}"
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    ws.merge_cells("A3:Z3")
    ws["A3"].value = (
        "Both assessments use identical exposure data and hazard likelihood scores. "
        "Differences reflect discipline-specific vulnerability and resilience weights. "
        "PH weights emphasize population health outcomes; EM weights emphasize critical infrastructure impacts."
    )
    ws["A3"].font = Font(name="Calibri", size=10, color="333333")
    ws["A3"].alignment = Alignment(wrap_text=True)

    ws.merge_cells("A4:Z4")
    ws["A4"].value = (
        "PH PHRAT weights: Natural Hazards 28%, Active Shooter 18%, Health 17%, Air Quality 12%, Extreme Heat 11%, Dam Failure 7%, Vector-Borne Disease 7%. "
        "EM PHRAT weights: Natural Hazards 32%, Active Shooter 13%, Extreme Heat 13%, Health 10%, Utilities 10%, Air Quality 8%, Dam Failure 8%, Vector-Borne Disease 6%."
    )
    ws["A4"].font = Font(name="Calibri", size=9, italic=True, color="666666")
    ws["A4"].alignment = Alignment(wrap_text=True)

    header_row = 6
    headers = ["Jurisdiction"]
    for _, domain_label in DOMAINS:
        headers.append(f"PH {domain_label}")
        headers.append(f"EM {domain_label}")
        headers.append(f"Difference")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    data_row = header_row + 1
    for row_data in rows:
        cell = ws.cell(row=data_row, column=1, value=row_data["name"])
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.border = THIN_BORDER

        col = 2
        for domain_key, _ in DOMAINS:
            ph_val = row_data["ph"].get(domain_key, 0)
            em_val = row_data["em"].get(domain_key, 0)
            diff = em_val - ph_val

            ph_cell = ws.cell(row=data_row, column=col, value=round(ph_val, 4))
            ph_cell.number_format = "0.0000"
            ph_cell.fill = PH_FILL
            ph_cell.border = THIN_BORDER
            ph_cell.alignment = Alignment(horizontal="center")

            em_cell = ws.cell(row=data_row, column=col + 1, value=round(em_val, 4))
            em_cell.number_format = "0.0000"
            em_cell.fill = EM_FILL
            em_cell.border = THIN_BORDER
            em_cell.alignment = Alignment(horizontal="center")

            diff_cell = ws.cell(row=data_row, column=col + 2, value=round(diff, 4))
            diff_cell.number_format = "+0.0000;-0.0000;0.0000"
            diff_cell.border = THIN_BORDER
            diff_cell.alignment = Alignment(horizontal="center")
            if diff > 0.005:
                diff_cell.font = DIFF_POS_FONT
            elif diff < -0.005:
                diff_cell.font = DIFF_NEG_FONT

            col += 3

        data_row += 1

    ws_notes = wb.create_sheet("Methodology Notes")
    notes = [
        ("CARA PH vs EM Risk Score Comparison — Methodology", ""),
        ("", ""),
        ("Formula", "PHRAT Score = sqrt(sum(w_i * Risk_i^2)) where p=2 (quadratic mean)"),
        ("", ""),
        ("Public Health Focus", "Population health outcomes, disease surveillance, vaccination coverage, vulnerable populations"),
        ("Emergency Management Focus", "Critical infrastructure, power grid vulnerability, housing/transportation, rural isolation, utilities"),
        ("", ""),
        ("Key Difference", "Both disciplines use the SAME exposure scores (hazard likelihood). They differ in how they weight vulnerability and resilience factors."),
        ("", ""),
        ("Positive Difference", "EM score is HIGHER than PH — infrastructure-focused assessment sees greater risk"),
        ("Negative Difference", "EM score is LOWER than PH — health-focused assessment sees greater risk"),
        ("", ""),
        ("PH Domain Weights", "Natural Hazards 28%, Active Shooter 18%, Health 17%, Air Quality 12%, Extreme Heat 11%, Dam Failure 7%, Vector-Borne Disease 7%"),
        ("EM Domain Weights", "Natural Hazards 32%, Active Shooter 13%, Extreme Heat 13%, Health 10%, Utilities 10%, Air Quality 8%, Dam Failure 8%, Vector-Borne Disease 6%"),
        ("", ""),
        ("Utilities Domain", "Supplementary in PH (not in PHRAT score); Primary in EM (10% weight)"),
        ("Dam Failure Domain", "Uses NID data, WI DNR dam inventory, and FEMA NRI flood data for county-level dam failure risk"),
        ("Vector-Borne Disease Domain", "Lyme disease, West Nile Virus, Anaplasmosis using WI DHS surveillance data and climate-adjusted models"),
        ("Data Sources", "FEMA NRI, OpenFEMA APIs, NOAA NCEI Storm Events, WI DHS, EPA AirNow, Census/ACS"),
    ]
    for row_idx, (label, value) in enumerate(notes, 1):
        cell_a = ws_notes.cell(row=row_idx, column=1, value=label)
        cell_b = ws_notes.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            cell_a.font = Font(name="Calibri", size=14, bold=True, color="1a237e")
        elif label and not value:
            cell_a.font = Font(name="Calibri", size=10, bold=True)
        else:
            cell_a.font = Font(name="Calibri", size=10, bold=True)
            cell_b.font = Font(name="Calibri", size=10)
    ws_notes.column_dimensions["A"].width = 25
    ws_notes.column_dimensions["B"].width = 100

    ws.freeze_panes = "B7"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, dir=tempfile.gettempdir())
    wb.save(tmp.name)
    tmp.close()
    logger.info(f"Generated EM comparison export: {tmp.name} with {len(rows)} jurisdictions")
    return tmp.name
