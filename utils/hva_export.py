import tempfile
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def generate_kaiser_hva_export(risk_data: dict) -> str:
    """
    Generate Kaiser Permanente HVA compatible Excel export
    
    Args:
        risk_data: Dictionary containing CARA risk assessment data
        
    Returns:
        str: Path to generated Excel file
    """
    try:
        # Create workbook
        wb = Workbook()
        
        # Create main HVA worksheet first
        ws_hva = wb.create_sheet("HVA Assessment")
        
        # Remove default sheet after creating our worksheet
        if len(wb.worksheets) > 1:
            # Find and remove the default "Sheet" worksheet
            for sheet in wb.worksheets:
                if sheet.title == "Sheet":
                    wb.remove(sheet)
                    break
        
        # Set up styles
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        regular_font = Font(size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Create header section
        ws_hva['A1'] = 'Kaiser Permanente Hazard Vulnerability Analysis'
        ws_hva['A1'].font = header_font
        ws_hva.merge_cells('A1:H1')
        ws_hva['A1'].alignment = center_align
        
        ws_hva['A3'] = f'HERC Region: {risk_data.get("location", "Unknown")}'
        ws_hva['A3'].font = subheader_font
        
        ws_hva['A4'] = f'Assessment Date: {datetime.now().strftime("%B %d, %Y")}'
        ws_hva['A4'].font = regular_font
        
        counties = risk_data.get("counties", [])
        if isinstance(counties, list):
            counties_str = ", ".join(str(county) for county in counties)
        else:
            counties_str = str(counties) if counties else "Unknown"
        ws_hva['A5'] = f'Counties: {counties_str}'
        ws_hva['A5'].font = regular_font
        
        # Create column headers
        headers = [
            'Hazard Type',
            'Probability (1-4)',
            'Human Impact (1-4)',  
            'Property Impact (1-4)',
            'Business Impact (1-4)',
            'Preparedness (1-4)',
            'Total Risk Score',
            'Priority Ranking'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_hva.cell(row=8, column=col)
            cell.value = header
            cell.font = subheader_font
            cell.alignment = center_align
            
        # Map CARA risk data to HVA format
        hva_hazards = [
            {
                'name': 'Flood',
                'cara_score': risk_data.get('natural_hazards', {}).get('flood', 0.0),
                'probability': _score_to_hva_scale(risk_data.get('natural_hazards', {}).get('flood', 0.0)),
                'human_impact': 3,
                'property_impact': 4,
                'business_impact': 3,
                'preparedness': 2
            },
            {
                'name': 'Tornado',
                'cara_score': risk_data.get('natural_hazards', {}).get('tornado', 0.0),
                'probability': _score_to_hva_scale(risk_data.get('natural_hazards', {}).get('tornado', 0.0)),
                'human_impact': 4,
                'property_impact': 4,
                'business_impact': 4,
                'preparedness': 2
            },
            {
                'name': 'Winter Storm',
                'cara_score': risk_data.get('natural_hazards', {}).get('winter_storm', 0.0),
                'probability': _score_to_hva_scale(risk_data.get('natural_hazards', {}).get('winter_storm', 0.0)),
                'human_impact': 2,
                'property_impact': 2,
                'business_impact': 3,
                'preparedness': 3
            },
            {
                'name': 'Thunderstorm',
                'cara_score': risk_data.get('natural_hazards', {}).get('thunderstorm', 0.0),
                'probability': _score_to_hva_scale(risk_data.get('natural_hazards', {}).get('thunderstorm', 0.0)),
                'human_impact': 2,
                'property_impact': 2,
                'business_impact': 2,
                'preparedness': 3
            },
            {
                'name': 'Infectious Disease Outbreak',
                'cara_score': risk_data.get('health_risk', 0.0),
                'probability': _score_to_hva_scale(risk_data.get('health_risk', 0.0)),
                'human_impact': 4,
                'property_impact': 1,
                'business_impact': 4,
                'preparedness': 2
            },
            {
                'name': 'Active Shooter/Violence',
                'cara_score': risk_data.get('active_shooter_risk', 0.0),
                'probability': _score_to_hva_scale(risk_data.get('active_shooter_risk', 0.0)),
                'human_impact': 4,
                'property_impact': 2,
                'business_impact': 3,
                'preparedness': 3
            },
            {
                'name': 'Extreme Heat',
                'cara_score': risk_data.get('extreme_heat_risk', 0.0),
                'probability': _score_to_hva_scale(risk_data.get('extreme_heat_risk', 0.0)),
                'human_impact': 3,
                'property_impact': 1,
                'business_impact': 2,
                'preparedness': 2
            },
            {
                'name': 'Air Quality Emergency',
                'cara_score': risk_data.get('air_quality_risk', 0.0),
                'probability': _score_to_hva_scale(risk_data.get('air_quality_risk', 0.0)),
                'human_impact': 3,
                'property_impact': 1,
                'business_impact': 2,
                'preparedness': 2
            },
            {
                'name': 'Utility Outage',
                'cara_score': risk_data.get('utilities_risk', 0.0),
                'probability': _score_to_hva_scale(risk_data.get('utilities_risk', 0.0)),
                'human_impact': 3,
                'property_impact': 2,
                'business_impact': 4,
                'preparedness': 2
            },
            {
                'name': 'Cybersecurity Incident',
                'cara_score': risk_data.get('cybersecurity_risk', 0.0),
                'probability': _score_to_hva_scale(risk_data.get('cybersecurity_risk', 0.0)),
                'human_impact': 2,
                'property_impact': 1,
                'business_impact': 4,
                'preparedness': 2
            }
        ]
        
        # Calculate risk scores and add data rows
        for i, hazard in enumerate(hva_hazards, 9):
            # Calculate total risk score (Probability × Human Impact × Business Impact / Preparedness)
            total_risk = (hazard['probability'] * hazard['human_impact'] * hazard['business_impact']) / max(1, hazard['preparedness'])
            hazard['total_risk'] = total_risk
            
            # Add row data
            ws_hva.cell(row=i, column=1, value=hazard['name'])
            ws_hva.cell(row=i, column=2, value=hazard['probability'])
            ws_hva.cell(row=i, column=3, value=hazard['human_impact'])
            ws_hva.cell(row=i, column=4, value=hazard['property_impact'])
            ws_hva.cell(row=i, column=5, value=hazard['business_impact'])
            ws_hva.cell(row=i, column=6, value=hazard['preparedness'])
            ws_hva.cell(row=i, column=7, value=round(total_risk, 2))
        
        # Sort hazards by risk score for priority ranking
        hva_hazards.sort(key=lambda x: x['total_risk'], reverse=True)
        
        # Add priority rankings
        for i, hazard in enumerate(hva_hazards):
            # Find the row for this hazard
            for row in range(9, 9 + len(hva_hazards)):
                if ws_hva.cell(row=row, column=1).value == hazard['name']:
                    ws_hva.cell(row=row, column=8, value=i + 1)
                    break
        
        # Auto-fit columns
        for column in ws_hva.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_hva.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary sheet
        ws_summary = wb.create_sheet("Risk Summary")
        
        ws_summary['A1'] = 'HERC Region Risk Summary'
        ws_summary['A1'].font = header_font
        
        ws_summary['A3'] = f'Region: {risk_data.get("location", "Unknown")}'
        ws_summary['A4'] = f'Total Risk Score: {risk_data.get("total_risk_score", 0.0):.2f}'
        ws_summary['A5'] = f'Assessment Date: {datetime.now().strftime("%B %d, %Y")}'
        
        ws_summary['A7'] = 'Top 5 Risk Priorities:'
        ws_summary['A7'].font = subheader_font
        
        for i, hazard in enumerate(hva_hazards[:5], 8):
            ws_summary[f'A{i}'] = f'{i-7}. {hazard["name"]} (Risk Score: {hazard["total_risk"]:.2f})'
        
        # Save to temporary file with unique timestamp
        temp_dir = tempfile.gettempdir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'HERC_Region_{risk_data.get("herc_id", "Unknown")}_HVA_Export_{timestamp}.xlsx'
        file_path = os.path.join(temp_dir, filename)
        
        wb.save(file_path)
        
        return file_path
        
    except Exception as e:
        raise Exception(f"Error generating Kaiser Permanente HVA export: {str(e)}")

def _score_to_hva_scale(cara_score: float) -> int:
    """
    Convert CARA risk score (0.0-1.0) to HVA scale (1-4)
    
    Args:
        cara_score: CARA risk score between 0.0 and 1.0
        
    Returns:
        int: HVA scale value between 1 and 4
    """
    if cara_score <= 0.25:
        return 1  # Low
    elif cara_score <= 0.5:
        return 2  # Medium-Low
    elif cara_score <= 0.75:
        return 3  # Medium-High
    else:
        return 4  # High